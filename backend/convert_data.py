import openpyxl
import json

FILE_EXCEL = "data_halte.xlsx"

wb = openpyxl.load_workbook(FILE_EXCEL)

# ---- 1. haltes.json (Sheet11: name, lat, lon) ----
ws = wb["Sheet11"]
haltes = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue  # skip header
    name, lat, lon = row[0], row[1], row[2]
    if name and lat and lon:
        haltes.append({
            "name": str(name).strip(),
            "lat": float(lat),
            "lon": float(lon)
        })

with open("haltes.json", "w", encoding="utf-8") as f:
    json.dump(haltes, f, ensure_ascii=False, indent=4)
print(f"haltes.json → {len(haltes)} halte")

# ---- 2. transit.json (Sheet3: halte, koridor) ----
ws = wb["Sheet3"]
transit = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue  # skip header
    halte, koridor = row[0], row[1]
    if halte and koridor:
        transit.append({
            "halte": str(halte).strip(),
            "koridor": str(koridor).strip()
        })

with open("transit.json", "w", encoding="utf-8") as f:
    json.dump(transit, f, ensure_ascii=False, indent=4)
print(f"transit.json → {len(transit)} entri")

# ---- 3. edges.json (Sheet10: from, to, waktu, koridor) ----
ws = wb["Sheet10"]
edges = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: continue  # skip header
    from_, to, waktu, koridor = row[0], row[1], row[2], row[3]
    if from_ and to and waktu and koridor:
        edges.append({
            "from": str(from_).strip(),
            "to": str(to).strip(),
            "waktu": int(waktu),
            "koridor": str(koridor).strip()
        })

with open("edges.json", "w", encoding="utf-8") as f:
    json.dump(edges, f, ensure_ascii=False, indent=4)
print(f"edges.json → {len(edges)} edges")

print("\nDone! 3 file JSON berhasil dibuat.")